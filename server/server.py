from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

BASE_DIR = os.getcwd()

FILE_NAME = os.path.join(BASE_DIR, "istoric_productie.xlsx")
FILE_STOC = os.path.join(BASE_DIR, "stoc_productie.xlsx")


# ================= STYLING =================

def style_header(ws):
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center


def auto_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2


def add_borders(ws, start_row, end_row):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = border
            cell.alignment = center


def merge_vertical(ws, col, start_row, end_row):
    if start_row != end_row:
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=end_row, end_column=col)
    ws.cell(row=start_row, column=col).alignment = Alignment(
        horizontal="center", vertical="center"
    )


# ================= INIT EXCEL PRODUCTIE =================

def init_excel():
    if not os.path.exists(FILE_NAME):

        wb = Workbook()
        ws = wb.active
        ws.title = "Productie"

        ws.append([
            "Data server",
            "Proiect",
            "Utilizator",
            "Meteo",
            "Orele de ploaie",
            "Lucrare",
            "Cantitate",
            "Utilaj",
            "Nr utilaje",
            "Personal",
            "Nr personal",
            "Observatii"
        ])

        style_header(ws)

        wb.save(FILE_NAME)


# ================= INIT EXCEL STOC =================

def init_stoc_excel():

    if not os.path.exists(FILE_STOC):

        wb = Workbook()
        ws = wb.active
        ws.title = "Stoc"

        ws.append([
            "ID",
            "Unealta",
            "Persoana",
            "Data predare"
        ])

        style_header(ws)

        wb.save(FILE_STOC)


# ================= SAVE PRODUCTIE =================

@app.route("/save", methods=["POST"])
def save_data():

    try:

        data = request.json

        init_excel()

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        start_row = ws.max_row + 1

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        project = data.get("project")
        username = data.get("username")
        meteo = data.get("meteo")
        observatii = data.get("observatii")

        ploaie_text = ""

        if data.get("ploaieInterval"):

            start = data["ploaieInterval"].get("start")
            stop = data["ploaieInterval"].get("stop")

            if start and stop:
                ploaie_text = f"{start} - {stop}"

        lucrari = data.get("lucrari", [])

        utilaje = [(k, v) for k, v in data.get("utilaje", {}).items() if v > 0]
        personal = [(k, v) for k, v in data.get("personal", {}).items() if v > 0]

        max_rows = max(len(lucrari), len(utilaje), len(personal), 1)

        for i in range(max_rows):

            lucrare = ""
            cantitate = ""

            if i < len(lucrari):

                l = lucrari[i]

                lucrare = l.get("descriere")

                val = l.get("cantitate")

                if isinstance(val, float) and val.is_integer():
                    val = int(val)

                cantitate = f"{val} {l.get('unitate')}"

            utilaj = ""
            nr_utilaj = ""

            if i < len(utilaje):

                utilaj = utilaje[i][0]
                nr_utilaj = utilaje[i][1]

            rol = ""
            nr_personal = ""

            if i < len(personal):

                rol = personal[i][0]
                nr_personal = personal[i][1]

            ws.append([
                now if i == 0 else "",
                project if i == 0 else "",
                username if i == 0 else "",
                meteo if i == 0 else "",
                ploaie_text if i == 0 else "",
                lucrare,
                cantitate,
                utilaj,
                nr_utilaj,
                rol,
                nr_personal,
                observatii if i == 0 else ""
            ])

        end_row = ws.max_row

        merge_vertical(ws, 1, start_row, end_row)
        merge_vertical(ws, 2, start_row, end_row)
        merge_vertical(ws, 3, start_row, end_row)
        merge_vertical(ws, 4, start_row, end_row)
        merge_vertical(ws, 5, start_row, end_row)
        merge_vertical(ws, 12, start_row, end_row)

        add_borders(ws, start_row, end_row)

        style_header(ws)

        auto_width(ws)

        wb.save(FILE_NAME)

        return jsonify({"status": "ok"})

    except PermissionError:

        return jsonify({
            "error": "Fisierul Excel este deschis. Inchide-l si incearca din nou."
        }), 500


# ================= SAVE STOC UNEALTA =================

@app.route("/save_stoc", methods=["POST"])
def save_stoc():

    data = request.json

    init_stoc_excel()

    wb = load_workbook(FILE_STOC)

    ws = wb.active

    new_id = ws.max_row

    ws.append([
        new_id,
        data.get("unealta"),
        data.get("nume"),
        datetime.now().isoformat()
    ])

    auto_width(ws)

    add_borders(ws, 2, ws.max_row)

    wb.save(FILE_STOC)

    return jsonify({"status": "saved"})


# ================= GET STOC =================

@app.route("/get_stoc", methods=["GET"])
def get_stoc():

    init_stoc_excel()

    wb = load_workbook(FILE_STOC)

    ws = wb.active

    tools = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        tools.append({
            "id": row[0],
            "unealta": row[1],
            "nume": row[2],
            "data": row[3]
        })

    return jsonify(tools)

# ================= DASHBOARD STOC =================

@app.route("/dashboard_stoc", methods=["GET"])
def dashboard_stoc():

    init_stoc_excel()

    wb = load_workbook(FILE_STOC)
    ws = wb.active

    statistica = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        unealta = row[1]

        if unealta:

            if unealta not in statistica:
                statistica[unealta] = 0

            statistica[unealta] += 1

    rezultat = []

    for k, v in statistica.items():

        rezultat.append({
            "unealta": k,
            "total": v
        })

    return jsonify(rezultat)


# ================= DELETE TOOL =================

@app.route("/delete_tool", methods=["POST"])
def delete_tool():

    data = request.json

    tool_id = data.get("id")

    init_stoc_excel()

    wb = load_workbook(FILE_STOC)

    ws = wb.active

    for row in range(2, ws.max_row + 1):

        if ws.cell(row=row, column=1).value == tool_id:

            ws.delete_rows(row)

            break

    wb.save(FILE_STOC)

    return jsonify({"status": "deleted"})


# ================= DOWNLOAD FILES =================

@app.route("/download_istoric", methods=["GET"])
def download_istoric():

    init_excel()

    return send_file(
        FILE_NAME,
        as_attachment=True,
        download_name="istoric_productie.xlsx"
    )


@app.route("/download_stoc", methods=["GET"])
def download_stoc():

    init_stoc_excel()

    return send_file(
        FILE_STOC,
        as_attachment=True,
        download_name="stoc_productie.xlsx"
    )


# ================= RUN =================

if __name__ == "__main__":

    app.run(host="0.0.0.0", port=5000, debug=True)
